import pandas as pd
from openpyxl import load_workbook

excel_path = r"E:\Augment\jixiaotongji\5349094916_20250921_20251007代运营固定费用账单_1759889236749.xlsx"

# 使用openpyxl读取原始数据
wb = load_workbook(excel_path)
ws = wb.active

print("前5行原始数据:")
for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
    print(f"第{i}行: {row}")

print("\n使用pandas读取:")
df = pd.read_excel(excel_path)
print(f"列名: {df.columns.tolist()}")
print(f"\n前3行数据:")
print(df.head(3))
